from datetime import datetime
from pathlib import Path
import sqlite3

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from audit import audit_log
from config import BACKUP_DIR, BACKUPS_SQLITE_DIR, EXCEL_PATH, LOGS_DIR, PROJECT_ROOT, REPORTES_PDF_DIR
from data import limpiar_cache_reportes, preparar_dataframe
from services import backup_service
from ui.formatting import dataframe_visible


def _archivo_mtime(path):
    ruta = Path(path)
    return ruta.stat().st_mtime if ruta.exists() else 0


@st.cache_data(show_spinner=False)
def _contar_registros_excel_cached(path_text, mtime):
    path = Path(path_text)
    if not path.exists():
        return 0

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return 0

    try:
        hoja = workbook.active
        return max((hoja.max_row or 0) - 1, 0)
    finally:
        workbook.close()


def contar_registros_excel(path):
    ruta = Path(path)
    return _contar_registros_excel_cached(str(ruta.resolve()), _archivo_mtime(ruta))


def resumen_contrato_columnas(integridad):
    fuentes = [
        ("SQLite", integridad.get("columnas_no_canonicas_sqlite", []), integridad.get("columnas_extra_sqlite", [])),
        ("Excel", integridad.get("columnas_no_canonicas_excel", []), integridad.get("columnas_extra_excel", [])),
    ]
    filas = []
    for fuente, no_canonicas, extras in fuentes:
        for item in no_canonicas:
            filas.append({
                "Fuente": fuente,
                "Tipo": "No canónica",
                "Columna": item.get("columna", ""),
                "Corrección sugerida": item.get("columna_canonica", ""),
            })
        for columna in extras:
            filas.append({
                "Fuente": fuente,
                "Tipo": "Extra",
                "Columna": columna,
                "Corrección sugerida": "Revisar si debe agregarse al contrato",
            })
    return pd.DataFrame(filas, columns=["Fuente", "Tipo", "Columna", "Corrección sugerida"])


def renderizar_estado_contrato_columnas(st_module, integridad):
    resumen = resumen_contrato_columnas(integridad)
    if resumen.empty:
        st_module.caption("Contrato de columnas: OK, sin encabezados fuera de estándar.")
        return resumen

    st_module.warning("Se detectaron columnas fuera del contrato oficial.")
    st_module.dataframe(dataframe_visible(resumen), width="stretch", hide_index=True)
    return resumen


def _formato_fecha(path):
    path = Path(path)
    if not path.exists():
        return "No disponible"
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")


def _contar_archivos(path, patron="*"):
    path = Path(path)
    if not path.exists():
        return 0
    return sum(1 for item in path.glob(patron) if item.is_file())


def _ultimo_archivo(path, patron="*"):
    path = Path(path)
    if not path.exists():
        return None
    archivos = [item for item in path.glob(patron) if item.is_file()]
    if not archivos:
        return None
    return max(archivos, key=lambda item: item.stat().st_mtime)


def _integridad_sqlite(db_path):
    path = Path(db_path)
    if not path.exists():
        return "No existe"
    try:
        with sqlite3.connect(path) as connection:
            return connection.execute("PRAGMA integrity_check").fetchone()[0]
    except Exception as exc:
        return f"Error: {exc}"


def calcular_estado_sistema(df_reportes):
    import db

    existe_db = db.DB_PATH.exists()
    existe_excel = EXCEL_PATH.exists()
    registros_excel = contar_registros_excel(EXCEL_PATH)
    registros_sqlite = db.contar_registros() if existe_db else 0
    duplicados_detectados = db.contar_duplicados_operacionales() if existe_db else 0
    integridad = backup_service.verificar_integridad(db_path=db.DB_PATH, excel_path=EXCEL_PATH)
    coincide = registros_excel == registros_sqlite
    contrato = resumen_contrato_columnas(integridad)
    ultimo_backup = _ultimo_archivo(BACKUP_DIR)
    ultimo_backup_sqlite = _ultimo_archivo(BACKUPS_SQLITE_DIR)
    ultimo_pdf = _ultimo_archivo(REPORTES_PDF_DIR, "*.pdf")

    problemas = []
    if not existe_db:
        problemas.append("No existe la base SQLite principal.")
    if not existe_excel:
        problemas.append("No existe el Excel operacional.")
    if existe_db and _integridad_sqlite(db.DB_PATH) != "ok":
        problemas.append("SQLite no supera PRAGMA integrity_check.")
    if existe_db and existe_excel and not coincide:
        problemas.append("SQLite y Excel no tienen la misma cantidad de registros.")
    if duplicados_detectados:
        problemas.append("Existen duplicados operacionales históricos.")
    if not contrato.empty:
        problemas.append("Hay columnas fuera del contrato oficial.")
    if not ultimo_backup:
        problemas.append("No hay respaldos manuales en backup/.")

    estado_general = "OK" if not problemas else "Revisar"

    return {
        "estado_general": estado_general,
        "problemas": problemas,
        "integridad": integridad,
        "sqlite_integrity_check": _integridad_sqlite(db.DB_PATH),
        "existe_db": existe_db,
        "existe_excel": existe_excel,
        "registros_app": len(df_reportes),
        "registros_sqlite": registros_sqlite,
        "registros_excel": registros_excel,
        "duplicados": duplicados_detectados,
        "pdf_generados": _contar_archivos(REPORTES_PDF_DIR, "*.pdf"),
        "respaldos_backup": _contar_archivos(BACKUP_DIR),
        "respaldos_sqlite": _contar_archivos(BACKUPS_SQLITE_DIR),
        "logs": _contar_archivos(LOGS_DIR),
        "ultimo_pdf": ultimo_pdf,
        "ultimo_backup": ultimo_backup,
        "ultimo_backup_sqlite": ultimo_backup_sqlite,
        "fecha_db": _formato_fecha(db.DB_PATH),
        "fecha_excel": _formato_fecha(EXCEL_PATH),
    }


def _tabla_archivos_operativos(estado):
    filas = [
        {
            "Elemento": "SQLite principal",
            "Estado": "Existe" if estado["existe_db"] else "No existe",
            "Detalle": estado["sqlite_integrity_check"],
            "Última modificación": estado["fecha_db"],
        },
        {
            "Elemento": "Excel operacional",
            "Estado": "Existe" if estado["existe_excel"] else "No existe",
            "Detalle": f"{estado['registros_excel']:,} registros",
            "Última modificación": estado["fecha_excel"],
        },
        {
            "Elemento": "Último PDF",
            "Estado": "Disponible" if estado["ultimo_pdf"] else "Sin PDF",
            "Detalle": estado["ultimo_pdf"].name if estado["ultimo_pdf"] else "No disponible",
            "Última modificación": _formato_fecha(estado["ultimo_pdf"]) if estado["ultimo_pdf"] else "No disponible",
        },
        {
            "Elemento": "Último respaldo",
            "Estado": "Disponible" if estado["ultimo_backup"] else "Sin respaldo",
            "Detalle": estado["ultimo_backup"].name if estado["ultimo_backup"] else "No disponible",
            "Última modificación": _formato_fecha(estado["ultimo_backup"]) if estado["ultimo_backup"] else "No disponible",
        },
    ]
    return pd.DataFrame(filas)


def mostrar_estado_datos(df_reportes):
    with st.expander("Estado del sistema operacional", expanded=True):
        try:
            import db

            estado = calcular_estado_sistema(df_reportes)
            integridad = estado["integridad"]

            st.caption(f"Ruta oficial del proyecto: {PROJECT_ROOT}")
            st.caption(f"Base SQLite: {db.DB_PATH}")
            st.caption(f"Excel de respaldo/exportación: {EXCEL_PATH}")

            col_a, col_b, col_c, col_d, col_e, col_f = st.columns(6)
            col_a.metric("Estado", estado["estado_general"])
            col_b.metric("Registros app", f"{estado['registros_app']:,}")
            col_c.metric("Registros SQLite", f"{estado['registros_sqlite']:,}")
            col_d.metric("Registros Excel", f"{estado['registros_excel']:,}")
            col_e.metric("PDFs", f"{estado['pdf_generados']:,}")
            col_f.metric("Duplicados", f"{estado['duplicados']:,}")

            if estado["problemas"]:
                for problema in estado["problemas"]:
                    st.warning(problema)
            else:
                st.success("Sistema operativo sin observaciones críticas.")

            st.dataframe(dataframe_visible(_tabla_archivos_operativos(estado)), width="stretch", hide_index=True)
            renderizar_estado_contrato_columnas(st, integridad)

            if not df_reportes.empty:
                columnas_ultimo = [
                    columna
                    for columna in ["Fecha turno", "Modelo equipo", "Número equipo", "Operador", "Turno", "Metros perforados", "Hora registro"]
                    if columna in df_reportes.columns
                ]
                st.caption("Último registro leído por la app")
                st.dataframe(dataframe_visible(df_reportes.tail(1)[columnas_ultimo]), width="stretch", hide_index=True)

            col_sync, col_export = st.columns(2)
            with col_sync:
                if st.button("Migrar/sincronizar Excel hacia SQLite"):
                    df_actual = preparar_dataframe(pd.read_excel(EXCEL_PATH, engine="openpyxl"))
                    registros = db.reemplazar_dataframe_reportes(df_actual, source="manual_excel_to_sqlite")
                    audit_log.registrar_respaldo_sqlite(
                        resultado="ok",
                        detalle={"origen": "sincronizacion_manual", "registros": registros},
                    )
                    limpiar_cache_reportes()
                    st.success(f"SQLite sincronizado correctamente: {registros:,} registros.")

            with col_export:
                if st.button("Exportar SQLite a Excel respaldo"):
                    df_sqlite = db.leer_registros()
                    if df_sqlite.empty:
                        st.warning("SQLite no tiene registros para exportar.")
                    else:
                        BACKUPS_SQLITE_DIR.mkdir(exist_ok=True)
                        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                        export_path = BACKUPS_SQLITE_DIR / f"respaldo_sqlite_exportado_{timestamp}.xlsx"
                        df_sqlite.to_excel(export_path, index=False)
                        st.success(f"SQLite exportado a {export_path.name}: {len(df_sqlite):,} registros.")
        except Exception as exc:
            st.warning(f"No se pudo verificar el estado del sistema: {exc}")


def mostrar_estado_respaldo_sqlite(df_excel):
    mostrar_estado_datos(df_excel)
