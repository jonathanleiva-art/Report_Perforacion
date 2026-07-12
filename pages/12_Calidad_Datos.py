from io import BytesIO
from pathlib import Path
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import app_perforacion as app
from ui.page_header import render_page_header
import db
from operators import etiqueta_operador
from services import data_quality_service
from ui.formatting import dataframe_visible, texto_visible


def _opciones(columna):
    try:
        return db.obtener_valores_distintos_columna(columna)
    except Exception:
        return []


def _normalizar_rango(rango):
    if isinstance(rango, tuple) and len(rango) == 2 and all(valor is not None for valor in rango):
        return rango[0], rango[1]
    return None, None


def _filtros_sidebar():
    with app.st.sidebar:
        app.st.header("Filtros de calidad")
        rango = app.st.date_input("Rango de fechas", value=None, format="DD/MM/YYYY", key="calidad_fecha")
        turno = app.st.multiselect(
            "Turno",
            _opciones("Turno"),
            default=_opciones("Turno"),
            format_func=texto_visible,
            key="calidad_turno",
        )
        equipo = app.st.multiselect(
            "Equipo",
            _opciones("Modelo equipo"),
            default=_opciones("Modelo equipo"),
            format_func=texto_visible,
            key="calidad_equipo",
        )
        operador = app.st.multiselect(
            "Operador",
            _opciones("Operador"),
            default=_opciones("Operador"),
            format_func=lambda valor: texto_visible(etiqueta_operador(valor)),
            key="calidad_operador",
        )

    fecha_desde, fecha_hasta = _normalizar_rango(rango)
    return {
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "turno": turno,
        "equipo": equipo,
        "operador": operador,
    }


def _formato_estado(estado):
    return {
        "excelente": ("#16A34A", "Excelente", "La calidad está bajo control."),
        "aceptable": ("#2563EB", "Aceptable", "La calidad es estable con observaciones puntuales."),
        "observado": ("#D97706", "Observado", "Existen inconsistencias que requieren seguimiento."),
        "critico": ("#DC2626", "Crítico", "La calidad requiere corrección prioritaria."),
    }.get(
        estado.get("estado", "observado"),
        ("#D97706", "Observado", "Existen inconsistencias que requieren seguimiento."),
    )


def _mostrar_metricas(resultado):
    estado_color, estado_titulo, estado_texto = _formato_estado(resultado["estado"])
    col1, col2, col3, col4, col5 = app.st.columns(5)
    col1.metric("Score calidad", f"{resultado['score']:.2f}")
    col2.metric("Analizados", f"{resultado['evaluacion']['total_registros']:,.0f}")
    col3.metric("Errores", f"{resultado['evaluacion']['errores']:,.0f}")
    col4.metric("Advertencias", f"{resultado['evaluacion']['advertencias']:,.0f}")
    col5.metric("Reglas no evaluadas", f"{resultado['evaluacion']['reglas_no_evaluadas']:,.0f}")
    app.st.markdown(
        f"""
        <div style="border-left:6px solid {estado_color}; padding:0.75rem 1rem; background:#f8fafc; margin:0.5rem 0 1rem 0;">
            <strong>{estado_titulo}</strong><br/>
            {estado_texto}<br/>
            <span style="color:#475569;">{resultado['estado']['mensaje']}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _ordenar_detalle(detalle):
    if detalle.empty or "Estado" not in detalle.columns:
        return detalle

    orden = {"ERROR": 0, "WARNING": 1, "NO_EVALUADA": 2}
    resultado = detalle.copy()
    resultado["_orden"] = resultado["Estado"].map(orden).fillna(3)
    columnas = [col for col in ["_orden", "fila", "Regla", "Estado"] if col in resultado.columns]
    return resultado.sort_values(columnas).drop(columns=["_orden"])


def generar_excel_calidad(score, estado, analizados, errores, advertencias,
                           reglas_no_evaluadas, recomendacion, penalizacion,
                           df_observaciones, df_reglas_no_evaluadas):
    wb = Workbook()

    COLOR_HEADER    = "1E3A5F"
    COLOR_SUBHEADER = "2E5F8A"
    COLOR_NARANJA   = "F97316"
    COLOR_VERDE     = "166534"
    COLOR_AMARILLO  = "854F0B"
    COLOR_ROJO      = "991B1B"
    COLOR_FILA_PAR  = "EFF6FF"
    COLOR_FILA_IMPAR = "FFFFFF"
    BLANCO          = "FFFFFF"
    GRIS_BORDE      = "CBD5E1"

    borde_fino = Border(
        left=Side(style='thin', color=GRIS_BORDE),
        right=Side(style='thin', color=GRIS_BORDE),
        top=Side(style='thin', color=GRIS_BORDE),
        bottom=Side(style='thin', color=GRIS_BORDE),
    )

    def header_fill(color):
        return PatternFill("solid", fgColor=color)

    def aplicar_tabla(ws, ref, nombre, estilo="TableStyleMedium2"):
        tabla = Table(displayName=nombre, ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name=estilo, showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tabla)

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_view.showGridLines = False
    for col, ancho in zip("ABCDEFGH", [28, 22, 18, 18, 22, 18, 18, 28]):
        ws1.column_dimensions[col].width = ancho

    ws1.merge_cells('A1:H1')
    c = ws1['A1']
    c.value = "REPORTE DE CALIDAD DE DATOS — PERFORACIÓN TEPSAC"
    c.font = Font(bold=True, size=14, color=BLANCO)
    c.fill = header_fill(COLOR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells('A2:H2')
    c = ws1['A2']
    c.value = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  Sistema de Gestión Operacional de Perforación"
    c.font = Font(size=10, color=BLANCO, italic=True)
    c.fill = header_fill(COLOR_SUBHEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 18
    ws1.row_dimensions[3].height = 10

    kpis = [
        ("Score calidad",  f"{score:.2f}",        COLOR_NARANJA),
        ("Estado",         estado,                 COLOR_VERDE if estado == "Excelente" else COLOR_AMARILLO),
        ("Analizados",     str(analizados),         COLOR_SUBHEADER),
        ("Errores",        str(errores),            COLOR_ROJO if errores > 0 else COLOR_VERDE),
        ("Advertencias",   str(advertencias),       COLOR_AMARILLO if advertencias > 0 else COLOR_VERDE),
        ("Reglas N/E",     str(reglas_no_evaluadas), COLOR_SUBHEADER),
        ("Recomendación",  recomendacion,           COLOR_SUBHEADER),
        ("Penalización",   f"{penalizacion:.4f}",   COLOR_SUBHEADER),
    ]
    for i, (etiq, val, color) in enumerate(kpis):
        col = chr(ord('A') + i)
        ce = ws1[f'{col}4']
        ce.value = etiq
        ce.font = Font(bold=True, size=9, color=BLANCO)
        ce.fill = header_fill(COLOR_HEADER)
        ce.alignment = Alignment(horizontal='center', vertical='center')
        ce.border = borde_fino
        ws1.row_dimensions[4].height = 18
        cv = ws1[f'{col}5']
        cv.value = val
        cv.font = Font(bold=True, size=13, color=color)
        cv.fill = header_fill("F8FAFC")
        cv.alignment = Alignment(horizontal='center', vertical='center')
        cv.border = borde_fino
        ws1.row_dimensions[5].height = 28

    ws1.row_dimensions[6].height = 10

    if df_observaciones is not None and len(df_observaciones) > 0:
        ws1.merge_cells('A7:H7')
        c = ws1['A7']
        c.value = "PRINCIPALES PROBLEMAS DETECTADOS"
        c.font = Font(bold=True, size=11, color=BLANCO)
        c.fill = header_fill(COLOR_NARANJA)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws1.row_dimensions[7].height = 22

        # Agrupa por Regla/Estado/Recomendación — columna Estado (no "Estado predominante")
        grp_cols = [c for c in ['Regla', 'Estado', 'Recomendación operacional'] if c in df_observaciones.columns]
        resumen_grp = (
            df_observaciones.groupby(grp_cols).size()
            .reset_index(name='Cantidad')
            .sort_values('Cantidad', ascending=False)
            .head(10)
        )
        headers_res = ['Regla', 'Cantidad', 'Estado', 'Recomendación operacional']
        # índices posicionales en resumen_grp (orden: grp_cols + Cantidad)
        pos_map = {col: idx for idx, col in enumerate(grp_cols + ['Cantidad'])}

        fila_h = 8
        for j, h in enumerate(headers_res):
            c = ws1.cell(row=fila_h, column=j + 1, value=h)
            c.font = Font(bold=True, size=10, color=BLANCO)
            c.fill = header_fill(COLOR_SUBHEADER)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = borde_fino
        ws1.row_dimensions[fila_h].height = 20

        for r_idx, row in enumerate(resumen_grp.itertuples(index=False)):
            fila = fila_h + 1 + r_idx
            vals = [
                row[pos_map.get('Regla', 0)],
                row[pos_map.get('Cantidad', len(grp_cols))],
                row[pos_map.get('Estado', 1)] if 'Estado' in pos_map else '',
                row[pos_map.get('Recomendación operacional', 2)] if 'Recomendación operacional' in pos_map else '',
            ]
            fill_color = COLOR_FILA_PAR if r_idx % 2 == 0 else COLOR_FILA_IMPAR
            for j, val in enumerate(vals):
                c = ws1.cell(row=fila, column=j + 1, value=val)
                c.fill = PatternFill("solid", fgColor=fill_color)
                c.border = borde_fino
                c.alignment = Alignment(vertical='center', wrap_text=True)
                if j == 1:
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.font = Font(bold=True, color=COLOR_ROJO if 'ERROR' in str(vals[2]) else COLOR_AMARILLO)
            ws1.row_dimensions[fila].height = 18

        ultimo_fila_res = fila_h + len(resumen_grp)
        aplicar_tabla(ws1, f"A{fila_h}:D{ultimo_fila_res}", "TablaProblemas", "TableStyleMedium9")

    # ── Hoja 2: Observaciones ────────────────────────────────────────────
    ws2 = wb.create_sheet("Observaciones")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:K1')
    c = ws2['A1']
    c.value = "DETALLE DE OBSERVACIONES"
    c.font = Font(bold=True, size=13, color=BLANCO)
    c.fill = header_fill(COLOR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    if df_observaciones is not None and len(df_observaciones) > 0:
        cols_obs = list(df_observaciones.columns)
        anchos = {
            'Regla': 30, 'Estado': 14, 'fila': 8, 'Fecha turno': 14, 'Turno': 10,
            'Modelo equipo': 18, 'Número equipo': 14, 'Operador': 18,
            'Mensaje': 45, 'Recomendación operacional': 40, 'Valor observado': 18,
        }
        for j, col in enumerate(cols_obs):
            ws2.column_dimensions[get_column_letter(j + 1)].width = anchos.get(col, 16)

        for j, col in enumerate(cols_obs):
            c = ws2.cell(row=2, column=j + 1, value=col)
            c.font = Font(bold=True, size=10, color=BLANCO)
            c.fill = header_fill(COLOR_SUBHEADER)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = borde_fino
        ws2.row_dimensions[2].height = 22

        for r_idx, row in enumerate(df_observaciones.itertuples(index=False)):
            fila = 3 + r_idx
            estado_fila = str(row[cols_obs.index('Estado')] if 'Estado' in cols_obs else '').upper()
            if 'ERROR' in estado_fila:
                row_color, font_color_estado = "FEE2E2", COLOR_ROJO
            elif 'WARNING' in estado_fila or 'ADVERTENCIA' in estado_fila:
                row_color, font_color_estado = "FEF3C7", COLOR_AMARILLO
            else:
                row_color = COLOR_FILA_PAR if r_idx % 2 == 0 else COLOR_FILA_IMPAR
                font_color_estado = COLOR_VERDE

            for j, col in enumerate(cols_obs):
                val = row[j]  # acceso posicional — seguro con cualquier nombre de columna
                c = ws2.cell(row=fila, column=j + 1, value=val)
                c.fill = PatternFill("solid", fgColor=row_color)
                c.border = borde_fino
                c.alignment = Alignment(vertical='center', wrap_text=True)
                if col == 'Estado':
                    c.font = Font(bold=True, color=font_color_estado)
            ws2.row_dimensions[fila].height = 16

        ultimo = 2 + len(df_observaciones)
        aplicar_tabla(ws2, f"A2:{get_column_letter(len(cols_obs))}{ultimo}", "TablaObservaciones", "TableStyleMedium2")

    # ── Hoja 3: Reglas no evaluadas ──────────────────────────────────────
    ws3 = wb.create_sheet("Reglas no evaluadas")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:D1')
    c = ws3['A1']
    c.value = "REGLAS NO EVALUADAS"
    c.font = Font(bold=True, size=13, color=BLANCO)
    c.fill = header_fill(COLOR_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    if df_reglas_no_evaluadas is not None and len(df_reglas_no_evaluadas) > 0:
        cols_r = list(df_reglas_no_evaluadas.columns)
        for j, col in enumerate(cols_r):
            ws3.column_dimensions[get_column_letter(j + 1)].width = 35
            c = ws3.cell(row=2, column=j + 1, value=col)
            c.font = Font(bold=True, size=10, color=BLANCO)
            c.fill = header_fill(COLOR_SUBHEADER)
            c.border = borde_fino
            c.alignment = Alignment(horizontal='center')
        ws3.row_dimensions[2].height = 20

        for r_idx, row in enumerate(df_reglas_no_evaluadas.itertuples(index=False)):
            fila = 3 + r_idx
            fill_color = COLOR_FILA_PAR if r_idx % 2 == 0 else COLOR_FILA_IMPAR
            for j in range(len(cols_r)):
                val = row[j]
                c = ws3.cell(row=fila, column=j + 1, value=val)
                c.fill = PatternFill("solid", fgColor=fill_color)
                c.border = borde_fino
                c.alignment = Alignment(vertical='center', wrap_text=True)
            ws3.row_dimensions[fila].height = 16

        ultimo_r = 2 + len(df_reglas_no_evaluadas)
        aplicar_tabla(ws3, f"A2:{get_column_letter(len(cols_r))}{ultimo_r}", "TablaReglasNE", "TableStyleMedium9")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def main():
    if not app.requerir_acceso():
        return
    render_page_header(app.st, 
        "Calidad de Datos",
        f"Diagnóstico defensivo de datos operacionales | Fuente oficial: {db.DB_PATH.name}",
    )

    filtros = _filtros_sidebar()
    df = db.consultar_historial_filtrado(
        fecha_desde=filtros["fecha_desde"],
        fecha_hasta=filtros["fecha_hasta"],
        turno=filtros["turno"],
        equipo=filtros["equipo"],
        operador=filtros["operador"],
    )

    resultado = data_quality_service.generar_resumen_ejecutivo_calidad(df)
    detalle = _ordenar_detalle(resultado["evaluacion"]["detalle"])
    reglas_no_evaluadas = detalle[detalle["Estado"].astype(str).eq("NO_EVALUADA")].copy() if not detalle.empty and "Estado" in detalle.columns else detalle.head(0).copy()

    _mostrar_metricas(resultado)

    app.st.subheader("Resumen ejecutivo")
    _r = resultado["resumen"].iloc[0] if not resultado["resumen"].empty else pd.Series(dtype=object)
    c1, c2 = app.st.columns([2, 1])
    with c1:
        app.st.info(f"Recomendación operacional: {texto_visible(resultado['recomendacion_operacional'])}")
    with c2:
        app.st.download_button(
            "Descargar reporte Excel",
            data=generar_excel_calidad(
                score=float(_r.get("Score calidad", 0)),
                estado=str(_r.get("Estado", "")),
                analizados=int(_r.get("Analizados", 0)),
                errores=int(_r.get("Errores", 0)),
                advertencias=int(_r.get("Advertencias", 0)),
                reglas_no_evaluadas=int(_r.get("Reglas no evaluadas", 0)),
                recomendacion=str(_r.get("Recomendación operacional", "")),
                penalizacion=float(_r.get("Penalización por registro", 0)),
                df_observaciones=detalle,
                df_reglas_no_evaluadas=reglas_no_evaluadas,
            ),
            file_name="reporte_calidad_datos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    if resultado["evaluacion"]["reglas_no_evaluadas"] > 0:
        app.st.warning(f"Hay {resultado['evaluacion']['reglas_no_evaluadas']} regla(s) no evaluada(s) por columna faltante.")

    col_a, col_b = app.st.columns([1, 1])
    with col_a:
        app.st.subheader("Principales 5 problemas detectados")
        if resultado["top_problemas"].empty:
            app.st.success("No se detectaron problemas relevantes en los filtros actuales.")
        else:
            app.st.dataframe(dataframe_visible(resultado["top_problemas"]), width="stretch", hide_index=True)

    with col_b:
        app.st.subheader("Registros críticos priorizados")
        if resultado["registros_criticos"].empty:
            app.st.success("No hay registros críticos priorizados para los filtros actuales.")
        else:
            app.st.dataframe(dataframe_visible(resultado["registros_criticos"]), width="stretch", hide_index=True)

    app.st.subheader("Observaciones")
    if detalle.empty:
        app.st.success("No se detectaron observaciones para los filtros actuales.")
    else:
        app.st.dataframe(
            dataframe_visible(detalle),
            width="stretch",
            hide_index=True,
            column_config={
                "Regla": app.st.column_config.TextColumn("Regla", pinned=True),
                "Estado": app.st.column_config.TextColumn("Estado", pinned=True),
                "Recomendación operacional": app.st.column_config.TextColumn("Recomendación operacional"),
            },
        )


main()

