import base64
from pathlib import Path
import re
import sys

import pandas as pd

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import app_perforacion as app
from ui.page_header import render_page_header
import db
from services import malla_avance_service
from services.malla_service import (
    TIPOS_POZO_MALLA_CONTROL,
    TIPOS_SECTOR_PERFORACION,
    ESTADOS_SECTOR_PERFORACION,
    listar_archivos_planos_malla,
    listar_administracion_planos_malla,
    detectar_duplicados_plano_malla,
    eliminar_plano_malla,
    listar_auditoria_planos_malla,
    registrar_auditoria_plano_malla,
    listar_pozos_malla_control,
    limpiar_avance_malla,
    registrar_archivo_plano_malla,
    registrar_pozo_malla_control,
    obtener_preview_plano_malla,
    resumen_malla_control,
    actualizar_pozo_malla_control,
    obtener_archivo_plano_malla,
    registrar_plan_perforacion,
    listar_planes_perforacion,
    registrar_sector_perforacion,
    listar_sectores_perforacion,
    resumen_plan_perforacion,
    actualizar_sector_perforacion,
    desactivar_sector_perforacion,
    leer_auditoria_sectores_perforacion,
)
from services import clasificacion_operacional_service
from services import enaex_pdf_extraction_service
from ui.formatting import dataframe_visible, texto_visible


COLOR_TIPO_POZO = {
    "Producción": "#EF4444",
    "Buffer": "#10B981",
    "Precorte": "#059669",
    "Otro": "#64748B",
}

COLOR_ESTADO_POZO = {
    "pendiente": "#94A3B8",
    "realizado": "#2563EB",
}


def _mostrar_carga_plano():
    with app.st.expander("Cargar plano de perforación", expanded=True):
        app.st.caption("Carga manual de un PDF. El archivo queda guardado en SQLite y en la carpeta local del módulo.")
        with app.st.form("form_carga_plano_visual", clear_on_submit=False):
            archivo = app.st.file_uploader("Plano PDF", type=["pdf"], key="plano_pdf_control")
            accion_duplicado = "Mantener ambos"
            duplicados = None
            if archivo is not None:
                duplicados = detectar_duplicados_plano_malla(nombre_archivo=getattr(archivo, "name", ""))
                if duplicados is not None and not duplicados.empty:
                    app.st.warning("⚠ Plano posiblemente duplicado.")
                    app.st.dataframe(dataframe_visible(duplicados), width="stretch", hide_index=True)
                    accion_duplicado = app.st.radio(
                        "Acción ante duplicado",
                        ["Cancelar carga", "Mantener ambos", "Reemplazar existente"],
                        horizontal=True,
                        key="accion_duplicado_plano",
                    )
            enviar = app.st.form_submit_button("Guardar plano", type="primary")
            if enviar:
                if accion_duplicado == "Cancelar carga":
                    app.st.info("Carga cancelada por posible duplicado.")
                    return
                if accion_duplicado == "Reemplazar existente" and duplicados is not None and not duplicados.empty:
                    for _, fila in duplicados.iterrows():
                        eliminar_plano_malla(int(fila["id"]), observacion="Reemplazo por carga de plano duplicado.")
                resultado = registrar_archivo_plano_malla(archivo)
                if resultado["ok"]:
                    if accion_duplicado == "Reemplazar existente":
                        registrar_auditoria_plano_malla(
                            "reemplazo",
                            archivo=getattr(archivo, "name", ""),
                            id_plano=resultado["archivo_id"],
                            observacion="Reemplazo de plano posiblemente duplicado.",
                        )
                    app.st.success(resultado["mensaje"])
                    app.st.session_state["plano_control_seleccionado"] = resultado["archivo_id"]
                    app.st.rerun()
                else:
                    app.st.error(resultado["mensaje"])


def _mostrar_planos_cargados():
    app.st.subheader("Planos cargados")
    planos = listar_archivos_planos_malla()
    if planos.empty:
        app.st.info("Todavía no hay planos cargados.")
        return

    columnas = [
        columna
        for columna in ["id", "nombre_archivo", "ruta_archivo", "tipo_archivo", "fecha_carga", "categoria"]
        if columna in planos.columns
    ]
    app.st.dataframe(dataframe_visible(planos[columnas]), width="stretch", hide_index=True)


def _mostrar_administracion_planos():
    app.st.subheader("Administración de Planos")
    planos = listar_administracion_planos_malla()
    if planos.empty:
        app.st.info("No hay planos para administrar.")
        return

    columnas = {
        "id": "ID",
        "nombre_archivo": "Nombre archivo",
        "fecha_carga": "Fecha carga",
        "ruta_archivo": "Ruta",
        "banco": "Banco",
        "fase": "Fase",
        "estado": "Estado",
        "cantidad_sectores": "Cantidad sectores",
    }
    visibles = [col for col in columnas if col in planos.columns]
    app.st.dataframe(dataframe_visible(planos[visibles].rename(columns=columnas)), width="stretch", hide_index=True)

    with app.st.expander("Eliminar PDF", expanded=False):
        opciones = {
            f"ID {int(fila['id'])} | {fila.get('nombre_archivo', '')} | {fila.get('estado', '')}": int(fila["id"])
            for _, fila in planos.iterrows()
        }
        seleccion = app.st.selectbox("Plano a eliminar", list(opciones.keys()), key="admin_plano_eliminar")
        archivo_id = opciones[seleccion]
        observacion = app.st.text_input("Observación auditoría", key=f"obs_eliminar_plano_{archivo_id}")
        confirmar = app.st.checkbox("Confirmo eliminar PDF, plan y sectores asociados", key=f"confirmar_eliminar_plano_{archivo_id}")
        if app.st.button("Eliminar PDF", key=f"btn_eliminar_plano_{archivo_id}", disabled=not confirmar):
            resultado = eliminar_plano_malla(archivo_id, observacion=observacion or "Eliminación desde administración de planos.")
            if resultado["ok"]:
                app.st.success(resultado["mensaje"])
                app.st.rerun()
            else:
                app.st.error(resultado["mensaje"])

    auditoria = listar_auditoria_planos_malla()
    if not auditoria.empty:
        with app.st.expander("Auditoría de planos", expanded=False):
            app.st.dataframe(dataframe_visible(auditoria), width="stretch", hide_index=True)


def _mostrar_vista_previa_plano(detalle_plano):
    app.st.subheader("Vista previa del plano")
    preview_path = obtener_preview_plano_malla(detalle_plano.get("id"))
    if preview_path and Path(preview_path).exists():
        app.st.image(str(preview_path), caption="Vista previa rasterizada del plano", width="stretch")
        app.st.caption(f"Previsualización PNG guardada en: {preview_path}")
    else:
        app.st.info("No se generó una vista previa PNG para el plano seleccionado.")

    ruta_plano = Path(detalle_plano.get("ruta_archivo", ""))
    if ruta_plano.suffix.lower() == ".pdf" and ruta_plano.exists():
        with app.st.expander("Ver PDF original", expanded=False):
            try:
                pdf_bytes = ruta_plano.read_bytes()
                pdf_b64 = base64.b64encode(pdf_bytes).decode("utf-8")
                app.st.components.v1.html(
                    f"""
                    <iframe
                        src="data:application/pdf;base64,{pdf_b64}"
                        width="100%"
                        height="760"
                        style="border: 1px solid #CBD5E1; border-radius: 12px; background: #FFFFFF;"
                    ></iframe>
                    """,
                    height=780,
                    scrolling=True,
                )
                app.st.caption("El PDF original queda como respaldo visual.")
            except Exception:
                app.st.info("No fue posible incrustar el PDF directamente; la imagen rasterizada queda como vista principal.")


def _normalizar_tipo_pozo(valor):
    texto = texto_visible(valor).strip()
    if texto.lower().startswith("prod"):
        return "Producción"
    if texto.lower().startswith("buff"):
        return "Buffer"
    if texto.lower().startswith("prec"):
        return "Precorte"
    return "Otro"


def _normalizar_estado(valor):
    texto = texto_visible(valor).strip().lower()
    return "realizado" if texto in {"realizado", "hecho", "ok", "1", "true"} else "pendiente"


def _parsear_lineas_pozos(texto):
    registros = []
    errores = []
    for indice, linea in enumerate(texto.splitlines(), start=1):
        linea = linea.strip()
        if not linea:
            continue
        partes = [parte.strip() for parte in re.split(r"[,\t;|]+", linea) if parte.strip()]
        if not partes:
            continue
        numero_pozo = partes[0]
        tipo_pozo = _normalizar_tipo_pozo(partes[1]) if len(partes) > 1 else "Otro"
        metros_planificados = 0.0
        if len(partes) > 2:
            try:
                metros_planificados = float(str(partes[2]).replace(",", "."))
            except ValueError:
                errores.append(f"Línea {indice}: metros planificados inválidos.")
        estado = _normalizar_estado(partes[3]) if len(partes) > 3 else "pendiente"
        observacion = partes[4] if len(partes) > 4 else ""
        registros.append(
            {
                "numero_pozo": numero_pozo,
                "tipo_pozo": tipo_pozo,
                "metros_planificados": metros_planificados,
                "estado": estado,
                "realizado": estado == "realizado",
                "observacion": observacion,
            }
        )
    return registros, errores


def _opciones_planos(planos):
    if planos.empty:
        return {}
    return {
        f"{fila['nombre_archivo']} | {fila['fecha_carga']}": int(fila["id"])
        for _, fila in planos.iterrows()
    }


def _mostrar_kpis_plan(plan_id):
    resumen = resumen_plan_perforacion(plan_id)
    cols = app.st.columns(6)
    cols[0].metric("Total sectores", int(resumen["total_sectores"]))
    cols[1].metric("Pozos planificados", f"{float(resumen['total_pozos_planificados']):,.0f}")
    cols[2].metric("Metros planificados", f"{float(resumen['total_metros_planificados']):,.2f}")
    cols[3].metric("Producción", f"{float(resumen['produccion_planificada']):,.2f}")
    cols[4].metric("Buffer", f"{float(resumen['buffer_planificado']):,.2f}")
    cols[5].metric("Precorte", f"{float(resumen['precorte_planificado']):,.2f}")


def _tabla_resumen_sectores(sectores):
    if sectores.empty:
        return sectores
    tabla = sectores.copy()
    tabla["Sector"] = tabla["identificador_sector"].fillna("").astype(str)
    columnas = {
        "fase": "Fase",
        "banco": "Banco",
        "Sector": "Sector",
        "tipo_sector": "Tipo",
        "malla": "Malla",
        "numero_precorte": "Precorte",
        "secuencia_tronadura": "Secuencia",
        "pozos_planificados": "Pozos planificados",
        "metros_planificados": "Metros planificados",
        "pasadura": "Pasadura",
        "diametro": "Diámetro",
        "estado": "Estado",
    }
    visibles = [col for col in columnas if col in tabla.columns]
    return tabla[visibles].rename(columns=columnas)


def _mostrar_edicion_sectores(plan_id):
    sectores_todos = listar_sectores_perforacion(plan_id=plan_id, incluir_inactivos=True)
    if sectores_todos.empty:
        return

    with app.st.expander("Editar o desactivar sector planificado", expanded=False):
        opciones = {
            f"ID {int(fila['id'])} | {fila.get('tipo_sector', '')} | {fila.get('identificador_sector', '')} | activo {int(fila.get('activo', 1) or 0)}": int(fila["id"])
            for _, fila in sectores_todos.iterrows()
        }
        etiqueta = app.st.selectbox("Sector a corregir", list(opciones.keys()), key=f"editar_sector_{plan_id}")
        sector_id = opciones[etiqueta]
        sector = sectores_todos[sectores_todos["id"].eq(sector_id)].iloc[0].to_dict()

        with app.st.form(f"form_editar_sector_{sector_id}", clear_on_submit=False):
            col1, col2 = app.st.columns(2)
            tipo_sector = col1.selectbox(
                "Tipo sector",
                TIPOS_SECTOR_PERFORACION,
                index=TIPOS_SECTOR_PERFORACION.index(sector.get("tipo_sector")) if sector.get("tipo_sector") in TIPOS_SECTOR_PERFORACION else 0,
                key=f"edit_tipo_sector_{sector_id}",
            )
            estado = col2.selectbox(
                "Estado",
                ESTADOS_SECTOR_PERFORACION,
                index=ESTADOS_SECTOR_PERFORACION.index(sector.get("estado")) if sector.get("estado") in ESTADOS_SECTOR_PERFORACION else 0,
                key=f"edit_estado_sector_{sector_id}",
            )
            identificador_sector = app.st.text_input("Identificador sector", value=str(sector.get("identificador_sector") or ""), key=f"edit_identificador_{sector_id}")
            col3, col4 = app.st.columns(2)
            malla = col3.text_input("Malla", value=str(sector.get("malla") or ""), key=f"edit_malla_{sector_id}")
            numero_precorte = col4.text_input("Número de precorte", value=str(sector.get("numero_precorte") or ""), key=f"edit_precorte_{sector_id}")
            secuencia_tronadura = app.st.text_input("Secuencia tronadura", value=str(sector.get("secuencia_tronadura") or ""), key=f"edit_secuencia_{sector_id}")
            col5, col6 = app.st.columns(2)
            pozos_planificados = col5.number_input("Pozos planificados", min_value=0.0, step=1.0, value=float(sector.get("pozos_planificados") or 0), key=f"edit_pozos_{sector_id}")
            metros_planificados = col6.number_input("Metros planificados", min_value=0.0, step=1.0, value=float(sector.get("metros_planificados") or 0), key=f"edit_metros_{sector_id}")
            col7, col8 = app.st.columns(2)
            pasadura = col7.text_input("Pasadura", value=str(sector.get("pasadura") or ""), key=f"edit_pasadura_{sector_id}")
            diametro = col8.text_input("Diámetro", value=str(sector.get("diametro") or ""), key=f"edit_diametro_{sector_id}")
            observacion = app.st.text_area("Observación sector", value=str(sector.get("observacion") or ""), key=f"edit_obs_{sector_id}")
            motivo = app.st.text_input("Motivo de corrección", key=f"edit_motivo_{sector_id}")
            guardar = app.st.form_submit_button("Guardar corrección")

        col_accion_1, col_accion_2 = app.st.columns(2)
        motivo_desactivar = col_accion_1.text_input("Motivo desactivación", key=f"motivo_desactivar_sector_{sector_id}")
        desactivar = col_accion_2.button("Desactivar sector", key=f"btn_desactivar_sector_{sector_id}")

        if guardar:
            resultado = actualizar_sector_perforacion(
                sector_id,
                {
                    "tipo_sector": tipo_sector,
                    "identificador_sector": identificador_sector,
                    "malla": malla,
                    "numero_precorte": numero_precorte,
                    "secuencia_tronadura": secuencia_tronadura,
                    "pozos_planificados": pozos_planificados,
                    "metros_planificados": metros_planificados,
                    "pasadura": pasadura,
                    "diametro": diametro,
                    "estado": estado,
                    "observacion": observacion,
                },
                motivo=motivo,
            )
            if resultado["ok"]:
                app.st.success(resultado["mensaje"])
                app.st.rerun()
            else:
                app.st.error(resultado["mensaje"])

        if desactivar:
            resultado = desactivar_sector_perforacion(sector_id, motivo=motivo_desactivar)
            if resultado["ok"]:
                app.st.success(resultado["mensaje"])
                app.st.rerun()
            else:
                app.st.error(resultado["mensaje"])

        auditoria = leer_auditoria_sectores_perforacion(sector_id=sector_id)
        if not auditoria.empty:
            app.st.dataframe(dataframe_visible(auditoria), width="stretch", hide_index=True)


def _mostrar_extraccion_enaex_pdf(archivo_pdf_id, detalle_plano):
    if not archivo_pdf_id or not detalle_plano:
        return

    ruta_pdf = detalle_plano.get("ruta_archivo", "")
    nombre_archivo = detalle_plano.get("nombre_archivo", "")
    if not ruta_pdf:
        return

    with app.st.expander("Vista previa editable desde PDF Enaex", expanded=True):
        extraer = app.st.button("Analizar PDF Enaex", key=f"extraer_enaex_{archivo_pdf_id}")
        estado_key = f"enaex_extraccion_{archivo_pdf_id}"
        if extraer or estado_key not in app.st.session_state:
            app.st.session_state[estado_key] = enaex_pdf_extraction_service.extraer_datos_enaex_desde_pdf(
                ruta_pdf,
                nombre_archivo=nombre_archivo,
            )

        extraccion = app.st.session_state.get(estado_key) or {}
        with app.st.expander("Diagnóstico extracción PDF", expanded=True):
            app.st.write(f"Ruta PDF utilizada: {ruta_pdf}")
            app.st.write(f"Cantidad de texto extraído: {int(extraccion.get('texto_len', len(extraccion.get('texto_extraido', '') or '')))}")
            app.st.write(f"Fase detectada: {extraccion.get('fase', '')}")
            app.st.write(f"Banco detectado: {extraccion.get('banco', '')}")
            app.st.write(f"Malla detectada: {extraccion.get('malla', '')}")
            app.st.write(f"Sectores detectados: {len(extraccion.get('sectores', []) or [])}")
            errores = extraccion.get("errores") or []
            if errores:
                app.st.warning(" | ".join(errores))
            else:
                app.st.success("Sin errores críticos de extracción.")

        if not extraccion.get("ok"):
            app.st.info(extraccion.get("mensaje", "Selecciona un PDF con texto embebido para extraer datos."))
            texto_extraido = extraccion.get("texto_extraido", "")
            if texto_extraido:
                app.st.text_area("Texto PDF extraído para análisis", value=texto_extraido, height=220)
            return

        app.st.caption(extraccion.get("mensaje", "Texto PDF extraído. Revisa y corrige antes de guardar."))
        with app.st.form(f"form_guardar_enaex_{archivo_pdf_id}", clear_on_submit=False):
            col1, col2 = app.st.columns(2)
            nombre_plan = col1.text_input("Nombre plano", value=extraccion.get("nombre_plan", ""), key=f"enaex_nombre_{archivo_pdf_id}")
            fecha_plan = col2.text_input("Fecha", value=extraccion.get("fecha_plan", ""), key=f"enaex_fecha_{archivo_pdf_id}")
            col3, col4, col5 = app.st.columns(3)
            fase = col3.text_input("Fase", value=extraccion.get("fase", ""), key=f"enaex_fase_{archivo_pdf_id}")
            banco = col4.text_input("Banco", value=extraccion.get("banco", ""), key=f"enaex_banco_{archivo_pdf_id}")
            malla = col5.text_input("Malla", value=extraccion.get("malla", ""), key=f"enaex_malla_{archivo_pdf_id}")
            observacion = app.st.text_area(
                "Observación",
                value="Plan creado desde extracción de texto PDF Enaex.",
                key=f"enaex_obs_{archivo_pdf_id}",
                height=80,
            )

            sectores_df = enaex_pdf_extraction_service.sectores_a_dataframe(extraccion.get("sectores", []))
            if sectores_df.empty:
                app.st.warning("No se detectaron sectores. Puedes mantener la carga manual.")
                sectores_editados = sectores_df
            else:
                sectores_editados = app.st.data_editor(
                    sectores_df,
                    width="stretch",
                    hide_index=True,
                    num_rows="dynamic",
                    key=f"enaex_sectores_editor_{archivo_pdf_id}",
                )

            guardar = app.st.form_submit_button("Guardar plan y sectores detectados", type="primary")

        if guardar:
            resultado_plan = registrar_plan_perforacion(
                {
                    "nombre_plan": nombre_plan,
                    "archivo_pdf_id": archivo_pdf_id,
                    "ruta_pdf": ruta_pdf,
                    "fase": fase,
                    "banco": banco,
                    "fecha_plan": fecha_plan,
                    "observacion": observacion,
                }
            )
            if not resultado_plan["ok"]:
                app.st.error(resultado_plan["mensaje"])
                return

            errores = []
            for _, sector in sectores_editados.iterrows():
                datos_sector = sector.to_dict()
                datos_sector["plan_id"] = resultado_plan["plan_id"]
                datos_sector["malla"] = datos_sector.get("malla") or (malla if datos_sector.get("tipo_sector") == "Producción" else "")
                datos_sector["identificador_sector"] = datos_sector.get("identificador_sector") or datos_sector.get("tipo_sector")
                resultado_sector = registrar_sector_perforacion(datos_sector)
                if not resultado_sector["ok"]:
                    errores.append(resultado_sector["mensaje"])

            app.st.session_state["plan_perforacion_seleccionado"] = resultado_plan["plan_id"]
            registrar_auditoria_plano_malla(
                "edición",
                archivo=nombre_archivo,
                id_plano=archivo_pdf_id,
                observacion=f"Plan {resultado_plan['plan_id']} guardado desde extracción Enaex.",
            )
            if errores:
                app.st.warning("Plan guardado, pero algunos sectores requieren corrección manual: " + " | ".join(errores))
            else:
                app.st.success("Plan y sectores Enaex guardados correctamente.")
            app.st.rerun()


def _mostrar_control_planes_sectores():
    app.st.subheader("Control operacional por banco, fase, malla y sectores")
    planos = listar_archivos_planos_malla()

    with app.st.expander("Información general del plan", expanded=True):
        opciones = _opciones_planos(planos)
        detalle = {}
        with app.st.form("form_plan_perforacion", clear_on_submit=False):
            if opciones:
                seleccion_plano = app.st.selectbox("Plano PDF cargado", list(opciones.keys()), key="plan_pdf_cargado")
                archivo_pdf_id = opciones[seleccion_plano]
                detalle = obtener_archivo_plano_malla(archivo_pdf_id) or {}
                ruta_pdf = detalle.get("ruta_archivo", "")
            else:
                app.st.info("Puedes crear un plan sin PDF y asociarlo después, o cargar primero un plano PDF.")
                archivo_pdf_id = None
                ruta_pdf = app.st.text_input("Ruta PDF", key="plan_ruta_pdf_manual")

            col1, col2 = app.st.columns(2)
            fase = col1.text_input("Fase", key="plan_fase")
            banco = col2.text_input("Banco", key="plan_banco")
            fecha_plan = app.st.date_input("Fecha del plano", key="plan_fecha")
            nombre_plan = app.st.text_input("Nombre del plan", key="plan_nombre")
            observacion = app.st.text_area("Observación", key="plan_observacion", height=90)
            enviar = app.st.form_submit_button("Guardar plan", type="primary")

        if enviar:
            resultado = registrar_plan_perforacion(
                {
                    "nombre_plan": nombre_plan,
                    "archivo_pdf_id": archivo_pdf_id,
                    "ruta_pdf": ruta_pdf,
                    "fase": fase,
                    "banco": banco,
                    "fecha_plan": fecha_plan.isoformat() if fecha_plan else "",
                    "observacion": observacion,
                }
            )
            if resultado["ok"]:
                app.st.success(resultado["mensaje"])
                app.st.session_state["plan_perforacion_seleccionado"] = resultado["plan_id"]
                app.st.rerun()
            else:
                app.st.error(resultado["mensaje"])

        if opciones:
            _mostrar_extraccion_enaex_pdf(archivo_pdf_id, detalle)

    planes = listar_planes_perforacion()
    if planes.empty:
        app.st.info("Aún no hay planes de perforación registrados.")
        return

    opciones_planes = {
        f"{fila['nombre_plan']} | Fase {fila.get('fase', '')} | Banco {fila.get('banco', '')} | {fila.get('fecha_plan', '')}": int(fila["id"])
        for _, fila in planes.iterrows()
    }
    ids = list(opciones_planes.values())
    seleccionado_actual = app.st.session_state.get("plan_perforacion_seleccionado")
    indice = ids.index(seleccionado_actual) if seleccionado_actual in ids else 0
    seleccion_plan = app.st.selectbox(
        "Plan activo para sectores",
        list(opciones_planes.keys()),
        index=indice,
        key="plan_sector_activo",
    )
    plan_id = opciones_planes[seleccion_plan]
    app.st.session_state["plan_perforacion_seleccionado"] = plan_id

    app.st.subheader("Sectores de perforación del plano")
    with app.st.form("form_sector_perforacion", clear_on_submit=False):
        col1, col2 = app.st.columns(2)
        tipo_sector = col1.selectbox("Tipo sector", TIPOS_SECTOR_PERFORACION, key="sector_tipo")
        estado = col2.selectbox("Estado", ESTADOS_SECTOR_PERFORACION, key="sector_estado")
        identificador_sector = app.st.text_input(
            "Identificador sector",
            key="sector_identificador",
            placeholder="Producción Malla 114, Buffer 1, Precorte 01, Borde 1",
        )
        col3, col4 = app.st.columns(2)
        malla = col3.text_input("Malla", key="sector_malla")
        numero_precorte = col4.text_input("Número de precorte", key="sector_precorte", placeholder="01, 02, 03, 04, 05")
        secuencia_tronadura = app.st.text_input("Secuencia tronadura", key="sector_secuencia")
        col5, col6 = app.st.columns(2)
        pozos_planificados = col5.number_input("Pozos planificados", min_value=0.0, step=1.0, key="sector_pozos")
        metros_planificados = col6.number_input("Metros planificados", min_value=0.0, step=1.0, key="sector_metros")
        observacion_sector = app.st.text_area("Observación sector", key="sector_observacion", height=80)
        guardar_sector = app.st.form_submit_button("Agregar sector", type="primary")

    if guardar_sector:
        resultado = registrar_sector_perforacion(
            {
                "plan_id": plan_id,
                "tipo_sector": tipo_sector,
                "identificador_sector": identificador_sector,
                "malla": malla,
                "numero_precorte": numero_precorte,
                "secuencia_tronadura": secuencia_tronadura,
                "pozos_planificados": pozos_planificados,
                "metros_planificados": metros_planificados,
                "estado": estado,
                "observacion": observacion_sector,
            }
        )
        if resultado["ok"]:
            app.st.success(resultado["mensaje"])
            app.st.rerun()
        else:
            app.st.error(resultado["mensaje"])

    _mostrar_kpis_plan(plan_id)
    sectores = listar_sectores_perforacion(plan_id=plan_id)
    if sectores.empty:
        app.st.info("Aún no hay sectores registrados para este plan.")
    else:
        app.st.dataframe(dataframe_visible(_tabla_resumen_sectores(sectores)), width="stretch", hide_index=True)
    _mostrar_edicion_sectores(plan_id)


def _tabla_avance_sectores(sectores):
    if sectores.empty:
        return sectores
    columnas = {
        "fase": "fase",
        "banco": "banco",
        "tipo_sector": "tipo_sector",
        "malla": "malla",
        "precorte": "precorte",
        "pozos_planificados": "pozos planificados",
        "pozos_perforados": "pozos perforados",
        "pozos_pendientes": "pozos pendientes",
        "avance_pozos_pct": "avance pozos %",
        "metros_planificados": "metros planificados",
        "metros_perforados": "metros perforados",
        "metros_pendientes": "metros pendientes",
        "avance_metros_pct": "avance metros %",
        "estado_avance": "estado avance",
        "semaforo": "semaforo",
    }
    visibles = [col for col in columnas if col in sectores.columns]
    return sectores[visibles].rename(columns=columnas)


def _tabla_registros_avance(registros):
    if registros.empty:
        return registros
    columnas = {
        "fecha": "fecha",
        "turno": "turno",
        "equipo": "equipo",
        "operador": "operador",
        "fase": "fase",
        "banco": "banco",
        "malla": "malla",
        "tipo_sector": "tipo_sector",
        "numero_precorte": "numero_precorte",
        "identificador_sector": "identificador_sector",
        "pozos_perforados": "pozos perforados",
        "metros_perforados": "metros perforados",
    }
    visibles = [col for col in columnas if col in registros.columns]
    return registros[visibles].rename(columns=columnas)


def _mostrar_panel_clasificacion_operacional():
    app.st.subheader("Corrección de clasificación operacional")
    resumen = clasificacion_operacional_service.resumen_clasificacion_operacional()
    cols = app.st.columns(5)
    cols[0].metric("Registros", resumen["total_registros"])
    cols[1].metric("Con tipo sector", resumen["con_tipo_sector"])
    cols[2].metric("Producción inferida", resumen["inferidos_produccion"])
    cols[3].metric("Sin clasificar", resumen["sin_clasificar"])
    cols[4].metric("Precorte sin número", resumen["precorte_sin_numero"])

    if resumen["sin_clasificar"] or resumen["precorte_sin_numero"] or resumen["otro_sin_identificador"]:
        app.st.warning("Hay registros con datos faltantes para calcular avance completo.")

    with app.st.expander("Datos que faltan para calcular avance completo", expanded=bool(resumen["sin_clasificar"])):
        faltantes = clasificacion_operacional_service.listar_registros_clasificacion(solo_sin_clasificar=True, limit=300)
        if faltantes.empty:
            app.st.success("No hay registros totalmente sin clasificar en el rango consultado.")
        else:
            app.st.dataframe(dataframe_visible(faltantes), width="stretch", hide_index=True)

    registros = clasificacion_operacional_service.listar_registros_clasificacion(limit=300)
    if registros.empty:
        app.st.info("No hay registros operacionales disponibles para corregir.")
        return

    with app.st.expander("Corregir tipo de sector en registro existente", expanded=False):
        opciones = {
            f"ID {int(fila['id'])} | {fila.get('Fecha turno', '')} | {fila.get('Número equipo', '')} | {fila.get('clasificacion_operacional', '')}": int(fila["id"])
            for _, fila in registros.iterrows()
        }
        etiqueta = app.st.selectbox("Registro operacional", list(opciones.keys()), key="registro_clasificacion_selector")
        registro_id = opciones[etiqueta]
        registro = registros[registros["id"].eq(registro_id)].iloc[0].to_dict()

        with app.st.form(f"form_clasificacion_registro_{registro_id}", clear_on_submit=False):
            tipo_actual = registro.get("tipo_sector") or ""
            tipo_sector = app.st.selectbox(
                "Tipo de sector",
                ["", *clasificacion_operacional_service.TIPOS_SECTOR],
                index=(["", *clasificacion_operacional_service.TIPOS_SECTOR].index(tipo_actual) if tipo_actual in ["", *clasificacion_operacional_service.TIPOS_SECTOR] else 0),
                key=f"clasif_tipo_{registro_id}",
            )
            col1, col2 = app.st.columns(2)
            numero_precorte = col1.text_input("Número de precorte", value=str(registro.get("numero_precorte") or ""), key=f"clasif_precorte_{registro_id}")
            identificador_sector = col2.text_input("Identificador sector", value=str(registro.get("identificador_sector") or ""), key=f"clasif_identificador_{registro_id}")
            motivo = app.st.text_input("Motivo de corrección", key=f"clasif_motivo_{registro_id}")
            guardar = app.st.form_submit_button("Guardar clasificación")

        if guardar:
            resultado = clasificacion_operacional_service.actualizar_clasificacion_registro(
                registro_id,
                tipo_sector,
                numero_precorte=numero_precorte,
                identificador_sector=identificador_sector,
                motivo=motivo,
            )
            if resultado["ok"]:
                app.st.success(resultado["mensaje"])
                app.st.rerun()
            else:
                app.st.error(resultado["mensaje"])


def _mostrar_avance_real_planificado():
    app.st.subheader("Avance real vs planificado")
    planes = listar_planes_perforacion()
    if planes.empty:
        app.st.info("Primero registra un plan de perforación para calcular avance real.")
        return

    opciones = {
        f"{fila['nombre_plan']} | Fase {fila.get('fase', '')} | Banco {fila.get('banco', '')} | {fila.get('fecha_plan', '')}": int(fila["id"])
        for _, fila in planes.iterrows()
    }
    plan_id_actual = app.st.session_state.get("plan_perforacion_seleccionado")
    ids = list(opciones.values())
    indice = ids.index(plan_id_actual) if plan_id_actual in ids else 0
    seleccion = app.st.selectbox(
        "Plan para calcular avance",
        list(opciones.keys()),
        index=indice,
        key="avance_plan_selector",
    )
    plan_id = opciones[seleccion]

    with malla_avance_service.db.conectar_db(malla_avance_service.db.DB_PATH) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        resumen = malla_avance_service.obtener_resumen_avance_plan(conn, plan_id)

    avance = resumen["avance"]
    cols = app.st.columns(6)
    cols[0].metric("Pozos planificados", f"{float(avance['pozos_planificados_total']):,.0f}")
    cols[1].metric("Pozos perforados", f"{float(avance['pozos_perforados_total']):,.0f}")
    cols[2].metric("Avance pozos", f"{float(avance['avance_pozos_pct']):,.2f}%")
    cols[3].metric("Metros planificados", f"{float(avance['metros_planificados_total']):,.2f}")
    cols[4].metric("Metros perforados", f"{float(avance['metros_perforados_total']):,.2f}")
    cols[5].metric("Avance metros", f"{float(avance['avance_metros_pct']):,.2f}%")

    sectores = resumen["sectores"]
    if sectores.empty:
        app.st.info("El plan seleccionado no tiene sectores registrados.")
    else:
        app.st.caption(
            "Producción cruza por malla; Precorte por número de precorte; Buffer por tipo de sector; "
            "Borde/Otro por identificador cuando aplica."
        )
        app.st.warning(
            "Los registros sin clasificación operacional se interpretan como Producción solo cuando poseen malla asociada."
        )
        app.st.dataframe(dataframe_visible(_tabla_avance_sectores(sectores)), width="stretch", hide_index=True)

    registros = resumen["registros"]
    app.st.subheader("Detalle de registros usados para el cálculo")
    if registros.empty:
        app.st.info("No hay registros reales asociados al plan y sectores de producción.")
    else:
        app.st.dataframe(dataframe_visible(_tabla_registros_avance(registros)), width="stretch", hide_index=True)


def _mostrar_registro_asistido():
    planos = listar_archivos_planos_malla()
    app.st.subheader("Registrar pozos de la malla")
    if planos.empty:
        app.st.info("Primero carga un plano PDF para asociar los pozos.")
        return

    opciones = {
        f"{fila['nombre_archivo']} | {fila['fecha_carga']} | {fila['tipo_archivo']}": int(fila["id"])
        for _, fila in planos.iterrows()
    }
    plano_id_actual = app.st.session_state.get("plano_control_seleccionado")
    ids = list(opciones.values())
    indice = ids.index(plano_id_actual) if plano_id_actual in ids else 0

    seleccion = app.st.selectbox(
        "Plano cargado",
        list(opciones.keys()),
        index=indice,
        key="plano_control_edicion",
    )
    plano_id = opciones[seleccion]
    detalle_plano = obtener_archivo_plano_malla(plano_id)
    if detalle_plano is None:
        app.st.info("No fue posible cargar el plano seleccionado.")
        return

    app.st.caption(f"Archivo: {detalle_plano.get('nombre_archivo', '')}")
    app.st.caption(f"Ruta: {detalle_plano.get('ruta_archivo', '')}")
    app.st.caption("El plano se carga como PDF y se mantiene como referencia visual/manual. No se usa OCR todavía.")
    _mostrar_vista_previa_plano(detalle_plano)

    with app.st.form("form_carga_pozos_malla", clear_on_submit=False):
        app.st.caption("Pega una lista de pozos. Formato sugerido por línea: numero_pozo,tipo_pozo,metros_planificados,estado,observacion")
        texto_pozos = app.st.text_area(
            "Lista de pozos",
            height=180,
            key="pozos_lista_control",
            placeholder="1,Producción,12.5,pendiente\n2,Buffer,10,realizado\n3,Precorte,8,pendiente",
        )
        enviar = app.st.form_submit_button("Cargar pozos", type="primary")
        if enviar:
            registros, errores = _parsear_lineas_pozos(texto_pozos)
            if not registros:
                app.st.error("No se detectaron pozos válidos para cargar.")
            else:
                for registro in registros:
                    registro["archivo_plano_id"] = plano_id
                    registrar_pozo_malla_control(registro)
                if errores:
                    app.st.warning(" | ".join(errores))
                app.st.success(f"Se cargaron {len(registros):,} pozos en la malla.")
                app.st.session_state["plano_control_seleccionado"] = plano_id
                app.st.rerun()

    pozos = listar_pozos_malla_control(archivo_plano_id=plano_id)
    if pozos.empty:
        app.st.info("Aún no hay pozos cargados para este plano.")
        return

    filtro_tipo = app.st.multiselect(
        "Filtrar por tipo de pozo",
        sorted([valor for valor in pozos["tipo_pozo"].dropna().astype(str).unique() if valor]),
        default=[],
        key="filtro_tipo_malla",
    )
    filtro_estado = app.st.multiselect(
        "Filtrar por estado",
        sorted([valor for valor in pozos["estado"].dropna().astype(str).unique() if valor]),
        default=[],
        key="filtro_estado_malla",
    )

    filtrado = pozos.copy()
    if filtro_tipo:
        filtrado = filtrado[filtrado["tipo_pozo"].astype(str).isin(filtro_tipo)]
    if filtro_estado:
        filtrado = filtrado[filtrado["estado"].astype(str).isin(filtro_estado)]

    resumen = resumen_malla_control(archivo_plano_id=plano_id)
    c1, c2, c3, c4 = app.st.columns(4)
    c1.metric("Pozos totales", int(resumen["pozos_totales"]))
    c2.metric("Pozos realizados", int(resumen["pozos_realizados"]))
    c3.metric("Pozos pendientes", int(resumen["pozos_pendientes"]))
    c4.metric("Avance", f"{float(resumen['porcentaje_avance']):,.2f}%")

    c5, c6 = app.st.columns(2)
    c5.metric("Metros planificados", f"{float(resumen['metros_planificados_totales']):,.2f}")
    c6.metric("Avance por metros", f"{float(resumen['avance_metros']):,.2f}%")

    app.st.caption(
        "La marca realizado/pendiente es manual y se puede corregir en cualquier momento. "
        "La visualización usa colores por tipo y estado."
    )

    editable = filtrado.copy()
    editable["realizado"] = editable["realizado"].fillna(0).astype(int).astype(bool)
    editable["estado"] = editable["estado"].fillna("pendiente").astype(str)
    editable["tipo_pozo"] = editable["tipo_pozo"].fillna("Otro").astype(str)
    editable["observacion"] = editable["observacion"].fillna("").astype(str)
    columnas_editor = [col for col in ["id", "numero_pozo", "tipo_pozo", "metros_planificados", "estado", "realizado", "observacion"] if col in editable.columns]
    editado = app.st.data_editor(
        dataframe_visible(editable[columnas_editor]),
        width="stretch",
        hide_index=True,
        column_config={
            "realizado": app.st.column_config.CheckboxColumn("Realizado"),
        },
        key="editor_pozos_malla_control",
    )
    if app.st.button("Guardar cambios en la malla", key="guardar_cambios_malla_control"):
        if not isinstance(editado, pd.DataFrame):
            app.st.error("No se pudo leer la tabla editada.")
        else:
            for _, fila in editado.iterrows():
                actualizar_pozo_malla_control(
                    int(fila["id"]),
                    {
                        "numero_pozo": fila.get("numero_pozo", ""),
                        "tipo_pozo": fila.get("tipo_pozo", "Otro"),
                        "metros_planificados": fila.get("metros_planificados", 0),
                        "estado": "realizado" if bool(fila.get("realizado")) else "pendiente",
                        "realizado": bool(fila.get("realizado")),
                        "observacion": fila.get("observacion", ""),
                    },
                )
            app.st.success("Cambios guardados correctamente.")
            app.st.rerun()

    app.st.subheader("Vista por colores")
    if filtrado.empty:
        app.st.info("No hay registros para mostrar con los filtros actuales.")
        return

    for _, fila in filtrado.iterrows():
        tipo = str(fila.get("tipo_pozo", "Otro"))
        estado = str(fila.get("estado", "pendiente"))
        fondo = COLOR_TIPO_POZO.get(tipo, "#64748B")
        opacidad = "1" if estado == "realizado" else "0.35"
        borde = "#1D4ED8" if bool(fila.get("realizado")) else "#94A3B8"
        app.st.markdown(
            f"""
            <div style="
                border: 2px solid {borde};
                border-radius: 12px;
                padding: 12px 14px;
                margin-bottom: 10px;
                background: linear-gradient(90deg, {fondo}{'CC' if estado == 'realizado' else '55'}, {fondo}22);
                color: #0F172A;
            ">
                <strong>Pozo {texto_visible(fila.get('numero_pozo', ''))}</strong><br>
                Tipo: {texto_visible(tipo)} | Estado: {texto_visible(estado)} | Realizado: {"Sí" if bool(fila.get("realizado")) else "No"}<br>
                Metros planificados: {float(fila.get("metros_planificados", 0) or 0):,.2f}<br>
                Observación: {texto_visible(fila.get("observacion", ""))}
            </div>
            """,
            unsafe_allow_html=True,
        )

    if any(str(valor).lower() == "realizado" for valor in filtrado["estado"].astype(str).tolist()):
        app.st.caption("Los pozos realizados aparecen con borde más fuerte.")


def _mostrar_futuro_ocr_ia():
    with app.st.expander("Futuro OCR / IA — Capacidades planificadas", expanded=False):
        app.st.markdown(
            """
            Futuras capacidades documentadas:
            - detectar automáticamente números de pozo desde PDF;
            - identificar colores del plano;
            - clasificar Producción, Buffer y Precorte;
            - comparar foto del operador contra el plano;
            - marcar avance directamente sobre imagen.
            """
        )


# ── Funciones auxiliares: Lector PDF ────────────────────────────────────────


def _contar_paginas_pdf(pdf_bytes):
    try:
        import fitz
        import io as _io
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        return doc.page_count
    except Exception:
        return "?"


def _clasificar_pozo_por_numero(numero_str):
    s = str(numero_str).upper().strip()
    if s.startswith("P"):
        return "Precorte"
    if s.startswith("A"):
        return "Auxiliar"
    try:
        n = int(re.sub(r"[^0-9]", "", s))
        if 1 <= n <= 99:
            return "Precorte"
        if 100 <= n <= 199:
            return "Buffer 1"
        if 200 <= n <= 299:
            return "Buffer 2"
        if 300 <= n <= 399:
            return "Producción 300"
        if 400 <= n <= 499:
            return "Producción 400"
        if 500 <= n <= 599:
            return "Producción 500"
        if n >= 600:
            return "Producción 600+"
    except Exception:
        pass
    return None


def _extraer_pozos_desde_texto(texto, _pdf_bytes=b""):
    patron = re.compile(r"\b([PA][-_]?\d{1,4}|\d{3,4})\b")
    pozos = []
    for linea in texto.split("\n"):
        linea_clean = linea.strip()
        if not linea_clean:
            continue
        for m in patron.findall(linea_clean):
            tipo = _clasificar_pozo_por_numero(m)
            if tipo:
                pozos.append({
                    "Número pozo": m,
                    "Tipo": tipo,
                    "Línea origen": linea_clean[:60],
                    "Estado": "Pendiente",
                })
    if not pozos:
        return None
    df = pd.DataFrame(pozos).drop_duplicates(subset=["Número pozo"])
    return df.sort_values("Tipo").reset_index(drop=True)


def _extraer_pozos_ocr(pdf_bytes):
    import fitz
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        texto_total = ""
        for page in doc:
            texto_total += page.get_text() + "\n"
            for block in page.get_text("blocks"):
                texto_total += block[4] + "\n"
        if len(texto_total.strip()) > 50:
            return _extraer_pozos_desde_texto(texto_total)
        app.st.warning(
            "PDF escaneado sin texto embebido. Se requiere OCR externo (pytesseract). "
            "Por ahora usa el ingreso manual."
        )
        return None
    except Exception as exc:
        app.st.error(f"Error procesando PDF: {exc}")
        return None


def _mostrar_tabla_pozos_clasificada(df):
    COLORES = {
        "Precorte":        "#DBEAFE",
        "Buffer 1":        "#D1FAE5",
        "Buffer 2":        "#A7F3D0",
        "Producción 300":  "#FEF3C7",
        "Producción 400":  "#FDE68A",
        "Producción 500":  "#FCA5A5",
        "Producción 600+": "#F87171",
        "Auxiliar":        "#E9D5FF",
    }
    for tipo in df["Tipo"].unique():
        df_tipo = df[df["Tipo"] == tipo]
        color = COLORES.get(tipo, "#F3F4F6")
        app.st.markdown(
            f"""<div style="background:{color}20; border-left:3px solid {color};
                    border-radius:0 8px 8px 0; padding:6px 12px; margin-bottom:4px;">
                <span style="font-weight:600; font-size:13px; color:{color};">{tipo}</span>
                <span style="font-size:12px; color:rgba(255,255,255,0.55); margin-left:8px;">
                    {len(df_tipo)} pozos
                </span></div>""",
            unsafe_allow_html=True,
        )
        numeros = " · ".join(df_tipo["Número pozo"].tolist())
        app.st.markdown(
            f"<p style='font-size:12px; color:rgba(255,255,255,0.72); margin:0 0 10px 12px;'>{numeros}</p>",
            unsafe_allow_html=True,
        )


def _mostrar_resumen_tipos(df):
    tipos = df["Tipo"].value_counts()
    total = len(df)
    cols = app.st.columns(min(len(tipos), 4))
    for i, (tipo, count) in enumerate(tipos.items()):
        with cols[i % len(cols)]:
            app.st.metric(tipo, int(count), f"{count / total * 100:.0f}% del total")


def _mostrar_ingreso_manual_pozos():
    app.st.markdown("##### Ingreso manual de pozos")
    app.st.caption("Pega números de pozos separados por coma, espacio o salto de línea")
    texto = app.st.text_area(
        "Números de pozos",
        height=100,
        placeholder="101, 102, 103\n201, 202\nP001, P002\n301, 302...",
        key="lector_manual_pozos",
    )
    if texto.strip():
        df = _extraer_pozos_desde_texto(texto)
        if df is not None and not df.empty:
            _mostrar_tabla_pozos_clasificada(df)
            _mostrar_resumen_tipos(df)


def _generar_excel_pozos(df):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    FILL_TIPOS = {
        "Precorte":        "DBEAFE",
        "Buffer 1":        "D1FAE5",
        "Buffer 2":        "A7F3D0",
        "Producción 300":  "FEF3C7",
        "Producción 400":  "FDE68A",
        "Producción 500":  "FCA5A5",
        "Producción 600+": "F87171",
        "Auxiliar":        "E9D5FF",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "PLANO DE PERFORACIÓN — POZOS CLASIFICADOS"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Número pozo", "Tipo", "Estado", "Línea origen"]
    widths = [14, 18, 14, 45]
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=j + 1, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5F8A")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j + 1)].width = widths[j]

    for r_idx, (_, row) in enumerate(df.iterrows()):
        fila = 3 + r_idx
        color = FILL_TIPOS.get(str(row.get("Tipo", "")), "FFFFFF")
        for j, col in enumerate(headers):
            c = ws.cell(row=fila, column=j + 1, value=row.get(col, ""))
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(vertical="center")
            ws.row_dimensions[fila].height = 16

    for tipo in df["Tipo"].unique():
        df_t = df[df["Tipo"] == tipo]
        ws_t = wb.create_sheet(title=tipo[:31])
        ws_t.sheet_view.showGridLines = False
        ws_t.merge_cells("A1:C1")
        c = ws_t["A1"]
        c.value = f"{tipo} — {len(df_t)} pozos"
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_t.row_dimensions[1].height = 26
        for j, h in enumerate(["Número pozo", "Estado", "Línea origen"]):
            c = ws_t.cell(row=2, column=j + 1, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2E5F8A")
            ws_t.column_dimensions[get_column_letter(j + 1)].width = [14, 14, 45][j]
        color = FILL_TIPOS.get(tipo, "FFFFFF")
        for r_idx, (_, row) in enumerate(df_t.iterrows()):
            fila = 3 + r_idx
            for j, col in enumerate(["Número pozo", "Estado", "Línea origen"]):
                c = ws_t.cell(row=fila, column=j + 1, value=row.get(col, ""))
                c.fill = PatternFill("solid", fgColor=color)
                ws_t.row_dimensions[fila].height = 16

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _mostrar_lector_pdf_pozos():
    from ui.components import section_header as _sh
    import fitz
    import pdfplumber
    import io as _io

    _sh(
        "Lector de plano PDF",
        "Extrae automáticamente los pozos del plano y los clasifica por tipo",
        kicker="Lectura PDF",
        st_module=app.st,
    )

    pdf_file = app.st.file_uploader(
        "Subir plano PDF",
        type=["pdf"],
        key="lector_pdf_upload",
        help="Planos con texto o escaneados — el sistema detecta automáticamente",
    )
    if not pdf_file:
        app.st.info("Sube un plano PDF para extraer la tabla de pozos.")
        _mostrar_ingreso_manual_pozos()
        _mostrar_futuro_ocr_ia()
        return

    pdf_bytes = pdf_file.read()

    tiene_texto = False
    texto_extraido = ""
    try:
        with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf_doc:
            for page in pdf_doc.pages:
                texto_extraido += (page.extract_text() or "") + "\n"
        tiene_texto = len(texto_extraido.strip()) > 100
    except Exception:
        tiene_texto = False

    col_i1, col_i2 = app.st.columns(2)
    col_i1.metric("Método de lectura", "Extracción texto" if tiene_texto else "OCR visual")
    col_i2.metric("Páginas", _contar_paginas_pdf(pdf_bytes))
    app.st.divider()

    with app.st.spinner("Analizando plano y extrayendo pozos..."):
        pozos_df = (
            _extraer_pozos_desde_texto(texto_extraido) if tiene_texto
            else _extraer_pozos_ocr(pdf_bytes)
        )

    if pozos_df is None or pozos_df.empty:
        app.st.warning("No se detectaron pozos automáticamente. Usa el ingreso manual abajo.")
        _mostrar_ingreso_manual_pozos()
        _mostrar_futuro_ocr_ia()
        return

    col_prev, col_tabla = app.st.columns(2)
    with col_prev:
        app.st.markdown("##### Vista previa del plano")
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            from PIL import Image
            app.st.image(
                Image.frombytes("RGB", [pix.width, pix.height], pix.samples),
                width="stretch",
            )
        except Exception as exc:
            app.st.error(f"No se pudo renderizar el PDF: {exc}")

    with col_tabla:
        app.st.markdown("##### Pozos detectados")
        _mostrar_tabla_pozos_clasificada(pozos_df)

    app.st.divider()
    app.st.markdown("##### Resumen por tipo de pozo")
    _mostrar_resumen_tipos(pozos_df)

    app.st.divider()
    col_e1, col_e2 = app.st.columns(2)
    with col_e1:
        app.st.download_button(
            "Descargar CSV",
            pozos_df.to_csv(index=False).encode("utf-8"),
            "pozos_plano.csv",
            "text/csv",
            key="lector_dl_csv",
        )
    with col_e2:
        app.st.download_button(
            "Descargar Excel",
            _generar_excel_pozos(pozos_df),
            "pozos_plano.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="lector_dl_xlsx",
        )

    app.st.divider()
    app.st.markdown("##### Guardar en plan existente")
    planes = listar_planes_perforacion()
    if not planes.empty:
        opciones_planes = {
            f"{fila['nombre_plan']} | Fase {fila.get('fase', '')} | Banco {fila.get('banco', '')}": int(fila["id"])
            for _, fila in planes.iterrows()
        }
        app.st.selectbox("Plan destino", list(opciones_planes.keys()), key="lector_plan_destino")
        if app.st.button("Guardar pozos en plan", type="primary", key="lector_guardar_pozos"):
            app.st.info(
                "Para guardar los pozos usa la sección 'Registrar pozos de la malla' "
                "en la pestaña Plan y sectores, con el plano ya seleccionado."
            )
    else:
        app.st.info("Crea primero un plan en la pestaña 'Plan y sectores' para guardar los pozos.")

    _mostrar_futuro_ocr_ia()


# ── main ─────────────────────────────────────────────────────────────────────


def main():
    if not app.requerir_acceso():
        return
    from ui.components import section_header as _sh
    render_page_header(app.st, "Gestión de Planos")
    _sh(
        "Gestión de Planos",
        "Carga, lectura y análisis de planos de perforación · Clasificación de pozos por tipo",
        kicker="Planos",
        st_module=app.st,
    )
    app.st.caption(f"Fuente oficial: SQLite | Base: {db.DB_PATH}")

    tab1, tab2, tab3, tab4, tab5 = app.st.tabs([
        "📂 Planos cargados",
        "📋 Lector de plano PDF",
        "📊 Plan y sectores",
        "📈 Avance real vs planificado",
        "⚙️ Administración",
    ])

    with tab1:
        _mostrar_carga_plano()
        _mostrar_planos_cargados()

    with tab2:
        _mostrar_lector_pdf_pozos()

    with tab3:
        _mostrar_control_planes_sectores()
        app.st.divider()
        _mostrar_panel_clasificacion_operacional()
        app.st.divider()
        _mostrar_registro_asistido()

    with tab4:
        _mostrar_avance_real_planificado()

    with tab5:
        _mostrar_administracion_planos()


main()


