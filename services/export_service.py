from datetime import datetime
import logging
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import BACKUP_DIR


LOGGER = logging.getLogger(__name__)


def respaldar_excel_actual(path):
    path = Path(path)
    if not path.exists():
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    destino = BACKUP_DIR / f"{path.stem}_pre_save_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
    shutil.copy2(path, destino)
    return destino


def exportar_reportes_excel(df, path):
    path = Path(path)
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active
    encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fuente_encabezado = Font(color="FFFFFF", bold=True)
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = encabezado
        cell.font = fuente_encabezado

    encabezados = {cell.value: cell.column for cell in ws[1]}
    if "Fecha turno" in encabezados:
        col_fecha = encabezados["Fecha turno"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_fecha).number_format = "dd-mm-yyyy"

    for column_cells in ws.columns:
        ancho = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(ancho + 4, 45)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(path)
    return path
