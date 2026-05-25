from datetime import datetime
import logging
from pathlib import Path
import shutil
from unicodedata import normalize

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from audit import audit_log
from schema import NUMERIC_COLUMNS, TEXT_TAG_COLUMNS, aliases_columnas
from config import BACKUP_DIR
from utils import EXCEL_PATH, HORAS_TURNO, limpiar_entero


LOGGER = logging.getLogger(__name__)

COLUMNAS_CORREGIDAS = aliases_columnas()

REEMPLAZOS_TEXTO = {
    "Día": "Día",
    "día": "día",
    "Mantención": "Mantención",
    "mantención": "mantención",
    "Operación": "Operación",
    "operación": "operación",
    "Perforación": "Perforación",
    "perforación": "perforación",
    "Código": "Código",
    "C?digo": "Código",
    "Número": "Número",
    "número": "número",
    "Área": "Área",
    "Área": "Área",
    "Petróleo": "Petróleo",
    "Horómetro": "Horómetro",
    "detención": "detención",
    "Detención": "Detención",
    "Condición": "Condición",
    "condición": "condición",
    "Avería": "Avería",
    "avería": "avería",
    "Colación": "Colación",
    "Utilización": "Utilización",
    "utilización": "utilización",
    "Geología": "Geología",
}


def reparar_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()
    if texto.lower() in ("", "nan", "none", "nat"):
        return ""

    for _ in range(2):
        corregido = texto
        for encoding in ("latin1", "cp1252"):
            try:
                corregido = texto.encode(encoding).decode("utf-8")
                break
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
        if corregido == texto:
            break
        texto = corregido

    for origen, destino in REEMPLAZOS_TEXTO.items():
        texto = texto.replace(origen, destino)
    return texto.replace("", "").strip()


def es_lista_valores(valor):
    return isinstance(valor, (list, tuple, set, dict))


def valor_a_texto(valor):
    if valor is None:
        return ""
    try:
        if not es_lista_valores(valor) and pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(valor, dict):
        valores = valor.values()
    elif isinstance(valor, (list, tuple, set)):
        valores = valor
    else:
        return reparar_texto(valor)

    partes = []
    for item in valores:
        texto = valor_a_texto(item)
        if texto and texto.lower() not in ("nan", "none", "nat"):
            partes.append(texto)
    return ",".join(partes)


def valor_a_numero(valor):
    if isinstance(valor, dict):
        valores = valor.values()
    elif isinstance(valor, (list, tuple, set)):
        valores = valor
    else:
        valores = [valor]

    numeros = pd.to_numeric(
        pd.Series([valor_a_texto(item) for item in valores]),
        errors="coerce",
    ).fillna(0)
    return numeros.sum()


def fusionar_columnas_duplicadas(df):
    duplicadas = df.columns[df.columns.duplicated()].unique()
    if len(duplicadas) == 0:
        return df

    resultado = df.loc[:, ~df.columns.duplicated()].copy()
    for columna in duplicadas:
        bloque = df.loc[:, df.columns == columna]
        if columna in NUMERIC_COLUMNS:
            resultado[columna] = bloque.apply(
                lambda fila: pd.to_numeric(
                    fila.map(valor_a_texto),
                    errors="coerce",
                ).fillna(0).sum(),
                axis=1,
            )
        else:
            resultado[columna] = bloque.apply(
                lambda fila: ", ".join(
                    dict.fromkeys(
                        texto.strip()
                        for valor in fila
                        for texto in valor_a_texto(valor).split(",")
                        if texto.strip()
                    )
                ),
                axis=1,
            )

    return resultado


def normalizar_columnas(df):
    df = df.rename(columns={
        str(col).strip(): COLUMNAS_CORREGIDAS.get(
            reparar_texto(str(col).strip()),
            COLUMNAS_CORREGIDAS.get(str(col).strip(), reparar_texto(str(col).strip())),
        )
        for col in df.columns
    })
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed:")]
    return fusionar_columnas_duplicadas(df)


def normalizar_columna_texto(df, columna, enteros=False):
    if columna not in df.columns:
        return

    df[columna] = df[columna].apply(valor_a_texto)
    if enteros:
        df[columna] = df[columna].apply(normalizar_lista_enteros)
    else:
        df[columna] = df[columna].fillna("").astype(str).str.strip()


def normalizar_columna_numerica(df, columna):
    if columna not in df.columns:
        return

    serie = df[columna]
    if isinstance(serie, pd.DataFrame):
        serie = serie.apply(
            lambda fila: pd.to_numeric(fila.map(valor_a_texto), errors="coerce").fillna(0).sum(),
            axis=1,
        )
    else:
        serie = serie.apply(valor_a_numero)

    df[columna] = (
        pd.to_numeric(serie, errors="coerce")
        .replace([float("inf"), -float("inf")], 0)
        .fillna(0)
    )


def normalizar_fecha_turno(df):
    if "Fecha turno" not in df.columns:
        return df

    fechas = pd.to_datetime(df["Fecha turno"], errors="coerce", dayfirst=True)
    fechas_iso = pd.to_datetime(df["Fecha turno"], errors="coerce", format="%Y-%m-%d")
    df["Fecha turno"] = fechas.fillna(fechas_iso).dt.date
    return df


def normalizar_tipo_detencion(valor):
    if valor is None:
        return ""

    partes = []
    for item in valor_a_texto(valor).split(","):
        texto = item.strip()
        clave = normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii").lower()
        if clave in ("mecania", "mecanica", "averia mecanica"):
            texto = "Avería mecánica"
        if clave in ("sin marcacion", "standby por falta de tajo/patio"):
            texto = "Standby por falta de tajo/Patio"
        if clave in ("agua", "abastecimiento agua", "relleno de agua"):
            texto = "Relleno de agua"
        if clave in ("mantencion", "mantencion programada"):
            texto = "Mantención Programada"
        if clave == "cambio de aceros":
            texto = "Cambio de aceros"
        if clave == "geologia":
            texto = "Geología"
        if texto:
            partes.append(texto)

    return ", ".join(dict.fromkeys(partes))


def respaldar_excel_actual(path):
    path = Path(path)
    if not path.exists():
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    destino = BACKUP_DIR / f"{path.stem}_pre_save_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
    shutil.copy2(path, destino)
    return destino


def normalizar_lista_enteros(valor):
    if valor is None:
        return ""

    partes = []
    for item in valor_a_texto(valor).split(","):
        limpio = limpiar_entero(item)
        if limpio:
            partes.append(limpio)

    return ", ".join(partes)


def preparar_dataframe(df):
    if df.empty:
        return df

    df = normalizar_columnas(df.copy()).dropna(how="all")

    if "Número equipo" in df.columns:
        df["Número equipo"] = df["Número equipo"].apply(limpiar_entero)

    for col in ["Operador", "Modelo equipo", "Turno"]:
        normalizar_columna_texto(df, col)

    for col in TEXT_TAG_COLUMNS:
        normalizar_columna_texto(df, col, enteros=col in ("Banco", "Fase"))

    for col in NUMERIC_COLUMNS:
        normalizar_columna_numerica(df, col)

    if "Tipo detención" in df.columns:
        df["Tipo detención"] = df["Tipo detención"].apply(normalizar_tipo_detencion)

    df = normalizar_fecha_turno(df)

    if {"Modelo equipo", "Número equipo"}.issubset(df.columns):
        es_pv271_9291 = (
            df["Modelo equipo"].astype(str).str.strip().str.upper().eq("PV271")
            & df["Número equipo"].astype(str).apply(limpiar_entero).eq("9291")
        )
        df = df[~es_pv271_9291].copy()
        df["Equipo"] = (df["Modelo equipo"].astype(str) + " " + df["Número equipo"].astype(str)).str.strip()

    return df.reset_index(drop=True)


def excel_mtime():
    path = Path(EXCEL_PATH)
    return path.stat().st_mtime if path.exists() else 0


@st.cache_data(show_spinner=False)
def leer_excel_cached(path_text, mtime):
    path = Path(path_text)
    if not path.exists():
        return pd.DataFrame()

    df = pd.read_excel(path, engine="openpyxl")
    return preparar_dataframe(df)


def leer_reportes():
    return leer_excel_cached(str(EXCEL_PATH), excel_mtime())


def crear_registro(datos):
    registro = {
        "Fecha": datetime.now().strftime("%d-%m-%Y"),
        "Hora registro": datetime.now().strftime("%H:%M:%S"),
        "Horas turno": HORAS_TURNO,
        **datos,
    }
    return preparar_dataframe(pd.DataFrame([registro]))


def guardar_reportes(df):
    path = Path(EXCEL_PATH)
    respaldo = respaldar_excel_actual(path)
    df = preparar_dataframe(df)
    if "Fecha turno" in df.columns:
        df = df.sort_values("Fecha turno", na_position="last")

    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active
    encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fuente_encabezado = Font(color="FFFFFF", bold=True)
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = encabezado
        cell.font = fuente_encabezado

    encabezados = {cell.value: cell.column for cell in ws[1]}
    if "Fecha turno" in encabezados:
        col_fecha = encabezados["Fecha turno"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_fecha).number_format = "dd-mm-yyyy"

    for column_cells in ws.columns:
        ancho = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(ancho + 4, 45)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(path)
    respaldar_reportes_sqlite(df)
    return path, respaldo


def respaldar_reportes_sqlite(df):
    try:
        import db

        registros = db.reemplazar_dataframe_reportes(df)
        audit_log.registrar_respaldo_sqlite(
            resultado="ok",
            detalle={"registros": registros},
        )
    except Exception as exc:
        audit_log.registrar_respaldo_sqlite(
            resultado="error",
            detalle=str(exc),
        )
        LOGGER.exception("No se pudo respaldar reportes en SQLite.")


def anexar_registro(registro):
    existente = leer_reportes()
    final = pd.concat([existente, registro], ignore_index=True) if not existente.empty else registro
    path, respaldo = guardar_reportes(final)
    leer_excel_cached.clear()
    return final, path, respaldo


def limpiar_cache_reportes():
    leer_excel_cached.clear()
